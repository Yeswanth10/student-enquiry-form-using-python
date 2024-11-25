from tkinter import *
from tkinter import messagebox
import os
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
import datetime 
import pymysql




################################################## Database connection###################################
connection=pymysql.connect(host="localhost",user="root",passwd="",database="Student_details")
cursor=connection.cursor()
# print("ok")


############################################### Delete ##################################################

def delete():
    status=messagebox.askyesno(title='Qestion',message='Do You want to Close')
    if status==True:
        window.destroy()
    else:
        messagebox.showwarning(title='warning',message='Back to Form')


########################################## Save #########################################################

def save():
    status=messagebox.askyesno(title='Qestion',message='Do You want to save')
    if status==True:
        d=Date.get()
        n=Name.get()
        m=mobile.get()
        a=alter.get()
        e=email.get()
        ad=address.get()
        c=course.get()
        b=batch.get()
        h=hcku.get()
        ef1=ef.get()
        co=contact.get()
        cu=counselor.get()
        f=fee.get()
        comm=comment.get()
        # try:
        #     er=e_r
        # except:
        #     messagebox.showerror("error","select Enquery (or) Register")
        
        
        if d=="" or n=="" or m=="" or a=="" or e=="" or ad=="" or c=="" or b=="" or h=="" or ef1=="" or co=="" or cu=="" or f=="" or comm=="":
            messagebox.showerror("error","Few data is missing!")
        else:
            file=openpyxl.load_workbook('student_data.xlsx')
            sheet=file.active
            sheet.cell(column=1,row=sheet.max_row+1,value=d)
            sheet.cell(column=2,row=sheet.max_row,value=n)
            sheet.cell(column=3,row=sheet.max_row,value=m)
            sheet.cell(column=4,row=sheet.max_row,value=a)
            sheet.cell(column=5,row=sheet.max_row,value=e)
            sheet.cell(column=6,row=sheet.max_row,value=ad)
            sheet.cell(column=7,row=sheet.max_row,value=c)
            sheet.cell(column=8,row=sheet.max_row,value=b)
            sheet.cell(column=9,row=sheet.max_row,value=h)
            sheet.cell(column=10,row=sheet.max_row,value=ef1)
            sheet.cell(column=11,row=sheet.max_row,value=co)
            sheet.cell(column=12,row=sheet.max_row,value=cu)
            sheet.cell(column=13,row=sheet.max_row,value=f)
            sheet.cell(column=14,row=sheet.max_row,value=comm)
            # sheet.cell(column=15,row=sheet.max_row,value=er)
            
            ########################### data insert into Database#############################
            if register.get() == 1 and enquiry.get() == 1:
                insert_reg_record = "INSERT INTO Stu_Registration(Date,Name,Mobile_Number,Alternate_Number,Email_Id,Address,Course_Interested,Batch_Preferred,How_You_Came_To_Know_us,Experience_Fresher,Contact_Person_From_Besant,Counselor,Fees) VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s');"%(Date.get(),Name.get(),mobile.get(),alter.get(),email.get(),address.get(),course.get(),batch.get(),hcku.get(),ef.get(),contact.get(),counselor.get(),fee.get())
                insert_Enquiry_record = "INSERT INTO Enquiry(Date,Name,Mobile_Number,Alternate_Number,Email_Id,Address,Course_Interested,Batch_Preferred,How_You_Came_To_Know_us,Experience_Fresher,Contact_Person_From_Besant,Counselor,Fees) VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s');"%(Date.get(),Name.get(),mobile.get(),alter.get(),email.get(),address.get(),course.get(),batch.get(),hcku.get(),ef.get(),contact.get(),counselor.get(),fee.get())
            elif register.get() == 1:
                    insert_reg_record = "INSERT INTO Stu_Registration(Date,Name,Mobile_Number,Alternate_Number,Email_Id,Address,Course_Interested,Batch_Preferred,How_You_Came_To_Know_us,Experience_Fresher,Contact_Person_From_Besant,Counselor,Fees) VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s');"%(Date.get(),Name.get(),mobile.get(),alter.get(),email.get(),address.get(),course.get(),batch.get(),hcku.get(),ef.get(),contact.get(),counselor.get(),fee.get())
            elif enquiry.get() == 1:
                    insert_Enquiry_record = "INSERT INTO stu_Enquiry(Date,Name,Mobile_Number,Alternate_Number,Email_Id,Address,Course_Interested,Batch_Preferred,How_You_Came_To_Know_us,Experience_Fresher,Contact_Person_From_Besant,Counselor,Fees) VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s');"%(Date.get(),Name.get(),mobile.get(),alter.get(),email.get(),address.get(),course.get(),batch.get(),hcku.get(),ef.get(),contact.get(),counselor.get(),fee.get())
            else :
                    print("okk")
                    messagebox.showerror("error",'please select enquery or register')
            if register.get() == 1:
                cursor.execute(insert_reg_record)
            if enquiry.get() == 1:
                cursor.execute(insert_Enquiry_record)
            if register.get() == 1 or enquiry.get() == 1:
                connection.commit()
            file.save(r'student_data.xlsx')
            
            clear()
            
################################## clear ########################         
def clear():
    Name.delete(0, END)
    mobile.delete(0, END)
    alter.delete(0, END)
    email.delete(0, END)
    address.delete(0, END)
    course.delete(0, END)
    batch.delete(0, END)
    hcku.delete(0, END)
    ef.delete(0, END)
    contact.delete(0, END)
    counselor.delete(0, END)
    fee.delete(0, END)
    comment.delete(0, END)



################################### WORKBOOK CREATION ############################
file=pathlib.Path('student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Date"
    sheet['B1']="Name"
    sheet['C1']="Mobile No"
    sheet['D1']="Alter mobile No"
    sheet['E1']="Email"
    sheet['F1']="Address"
    sheet['G1']="Course intrested"
    sheet['H1']="Batch prefered"
    sheet['I1']="How u know us"
    sheet['J1']="Exprience r fresher"
    sheet['K1']="Contact person "
    sheet['L1']="counselor"
    sheet['M1']="Fee"
    sheet['N1']="Comment"
    sheet['O1']="Enqury/Register"    
    file.save('student_data.xlsx')
    
    
    
if __name__=="__main__" :
    window=Tk()
    window.title("Besant Technology")
    window.geometry("850x500+0+0")
    window.configure(bg="lightblue")
    Label(window,text="Besant Technology Enquiry Form",font="bold",bg="lightblue",fg="red",).grid(row=0,column=1,padx=20,pady=20)
    Date=Label(window,text="Date:",bg="lightblue").grid(row=1,column=0,sticky=W)
    # Entry(window,textvariable=search,width=25,bd=2).grid(row=1,column=3)
    Name=Label(window,text="Name:",bg="lightblue").grid(row=2,column=0,sticky=W)
    mobile=Label(window,text="Mobile Num:",bg="lightblue").grid(row=3,column=0,sticky=W)
    alter=Label(window,text="Alternate Num:",bg="lightblue").grid(row=4,column=0,sticky=W)
    email=Label(window,text="Email:",bg="lightblue").grid(row=5,column=0,sticky=W)
    adress=Label(window,text="Address",bg="lightblue").grid(row=6,column=0,sticky=W)
    course=Label(window,text="Course Intrested:",bg="lightblue").grid(row=7,column=0,sticky=W)
    batch=Label(window,text="Batch Prefered:",bg="lightblue").grid(row=8,column=0,sticky=W)
    hcku=Label(window,text="How You Come To Know Us:",bg="lightblue").grid(row=9,column=0,sticky=W)
    ef=Label(window,text="Are You Exprience or Fresher:",bg="lightblue").grid(row=10,column=0,sticky=W)
    contact=Label(window,text="Contact Person From Besant Technology:",bg="lightblue").grid(row=11,column=0,sticky=W)
    counselor=Label(window,text="Counselor:",bg="lightblue").grid(row=12,column=0,sticky=W)
    fee=Label(window,text="Fee:",bg="lightblue").grid(row=13,column=0,sticky=W)
    comment=Label(window,text="Comment:",bg="lightblue").grid(row=14,column=0,sticky=W)
    date=StringVar()
    today=datetime.date.today()
    d1=today.strftime("%d/%m/%y")
    Date=Entry(window,width=50,textvariable=date)
    Date.grid(row=1,column=1)
    date.set(d1)
    Name=Entry(window,width=50)
    Name.grid(row=2,column=1)
    mobile=Entry(window,width=50)
    mobile.grid(row=3,column=1)
    alter=Entry(window,width=50)
    alter.grid(row=4,column=1)
    email=Entry(window,width=50)
    email.grid(row=5,column=1)
    address=Entry(window,width=50)
    address.grid(row=6,column=1)
    course=Entry(window,width=50)
    course.grid(row=7,column=1)
    batch=Entry(window,width=50)
    batch.grid(row=8,column=1)
    hcku=Entry(window,width=50)
    hcku.grid(row=9,column=1)
    ef=Entry(window,width=50)
    ef.grid(row=10,column=1)
    contact=Entry(window,width=50)
    contact.grid(row=11,column=1)
    counselor=Entry(window,width=50)
    counselor.grid(row=12,column=1)
    fee=Entry(window,width=50)
    fee.grid(row=13,column=1)
    comment=Entry(window,width=50)
    comment.grid(row=14,column=1)


    #####################################bottons##############################
    enquiry=IntVar()
    register=IntVar()
    e_r=Checkbutton(window,text="Enquiry",variable=enquiry,bg="lightblue").grid(row=15,column=1,sticky=W,padx=10,pady=10)
    e_r=Checkbutton(window,text="Registration",variable=register,bg="lightblue").grid(row=15,column=1,sticky=N,padx=10,pady=10)
    save_button=Button(window,text="Submit",bg="green",command=save).grid(row=16,column=1,sticky=W,padx=10,pady=10)
    delete_button=Button(window,text="Delete",bg="red",command=delete).grid(row=16,column=1,sticky=N,padx=10,pady=10)
    # s_btn=Button(window,text="Search",width=10,command=search).grid(row=1,column=5,padx=5)
    my_button=Button(window,text="clear",bg="lightgray",command=clear).grid(row=16,column=1,sticky=E,padx=10,pady=10)
    # window2=Tk()
    # window.geometry("200x200")
    # window.configure(bg="white")
    # my_button1=Button(window2,text="YES",fg="black")
    # my_button1=Button(window2,text="NO",fg="black")


    window.mainloop()


